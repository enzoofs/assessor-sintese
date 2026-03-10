import json
import openpyxl
import os
import re

PRICE_DIR = 'listas_precos'
all_products = []

def clean_code(code):
    if code is None:
        return None
    return str(code).strip().upper()

def clean_price(price):
    if price is None:
        return None
    if isinstance(price, (int, float)):
        return round(float(price), 2)
    s = str(price).replace(',', '').replace('R$', '').replace('$', '').strip()
    try:
        return round(float(s), 2)
    except:
        return None

# ===== 1. HiMedia - EPL main =====
print('Loading HiMedia EPL main...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'EPL-USD-2025-updated-2025.xlsx'), read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    if 'not for export' in sheet_name.lower():
        continue
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=5, values_only=True):
        code = clean_code(row[0])
        desc = str(row[1]).strip() if row[1] else ''
        price = clean_price(row[2])
        if code and price and price > 0:
            all_products.append({
                'code': code,
                'description': desc,
                'price': price,
                'currency': 'USD',
                'brand': 'HIMEDIA',
                'source': 'EPL-USD-2025'
            })
wb.close()
print(f'  HiMedia EPL: {sum(1 for p in all_products if p["source"]=="EPL-USD-2025")} products')

# ===== 2. HiMedia - Growth Media G-series =====
print('Loading HiMedia G-series...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'Growth Media G-series prices NET FOB US$ to Sintese - Brazil-03072025.xlsx'), read_only=True, data_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=4, values_only=True):
    code = clean_code(row[0])
    desc = str(row[1]).strip() if row[1] else ''
    price = clean_price(row[2])
    if code and price and price > 0:
        if not any(p['code'] == code and p['brand'] == 'HIMEDIA' for p in all_products):
            all_products.append({
                'code': code,
                'description': desc,
                'price': price,
                'currency': 'USD',
                'brand': 'HIMEDIA',
                'source': 'HiMedia-G-series'
            })
wb.close()

# ===== 3. HiMedia - Updated MB prices =====
print('Loading HiMedia MB prices...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'Updated MB prices - EPL-USD-2025-NET FOB US$-26062025.xlsx'), read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    if 'not for export' in sheet_name.lower():
        continue
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=5, values_only=True):
        code = clean_code(row[0])
        desc = str(row[1]).strip() if row[1] else ''
        price = clean_price(row[2])
        if code and price and price > 0:
            existing = [p for p in all_products if p['code'] == code and p['brand'] == 'HIMEDIA']
            if existing:
                existing[0]['price'] = price
                existing[0]['source'] = 'HiMedia-MB-updated'
            else:
                all_products.append({
                    'code': code,
                    'description': desc,
                    'price': price,
                    'currency': 'USD',
                    'brand': 'HIMEDIA',
                    'source': 'HiMedia-MB-updated'
                })
wb.close()

# ===== 4. HiMedia - Instruments =====
print('Loading HiMedia Instruments...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'Instrument_price-2025-26.xlsx'), read_only=True, data_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=2, values_only=True):
    code = clean_code(row[0])
    desc = str(row[1]).strip() if row[1] else ''
    price = clean_price(row[2])
    if code and price and price > 0:
        existing = [p for p in all_products if p['code'] == code and p['brand'] == 'HIMEDIA']
        if existing:
            existing[0]['price'] = price
        else:
            all_products.append({
                'code': code,
                'description': desc,
                'price': price,
                'currency': 'USD',
                'brand': 'HIMEDIA',
                'source': 'HiMedia-Instruments'
            })
wb.close()
himedia_count = sum(1 for p in all_products if p['brand'] == 'HIMEDIA')
print(f'  Total HiMedia: {himedia_count} products')

# ===== 5. Benchmark Scientific =====
print('Loading Benchmark...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'Benchmark 2025 Master INTERNATIONAL Distributor Price List Final 10.15.24.xlsx'), read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=16, values_only=True):
        if len(row) < 4:
            continue
        code = clean_code(row[0])
        desc = str(row[1]).strip() if row[1] else ''
        price = clean_price(row[3])
        if code and price and price > 0 and not code.startswith('BENCHMARK'):
            brand = 'ACCURIS' if 'accuris' in desc.lower() else 'BENCHMARK'
            all_products.append({
                'code': code,
                'description': desc,
                'price': price,
                'currency': 'USD',
                'brand': brand,
                'source': 'Benchmark-2025'
            })
wb.close()
print(f'  Benchmark/Accuris: {sum(1 for p in all_products if p["brand"] in ["BENCHMARK","ACCURIS"])} products')

# ===== 6. F1001-F1008 (Accuris) =====
print('Loading Accuris F1001-F1008...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'F1001_F1008 International Distributor Pricing.xlsx'), read_only=True, data_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=4, values_only=True):
    if len(row) < 4:
        continue
    code = clean_code(row[0])
    desc = str(row[1]).strip() if row[1] else ''
    price = clean_price(row[3])
    if code and price and price > 0:
        all_products.append({
            'code': code,
            'description': desc,
            'price': price,
            'currency': 'USD',
            'brand': 'ACCURIS',
            'source': 'Accuris-F1001'
        })
wb.close()

# ===== 7. Hermle =====
print('Loading Hermle...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'Hermle 2025 INTL Price List, final 10_15_24.xlsx'), read_only=True, data_only=True)
ws = wb.active
for row in ws.iter_rows(min_row=14, values_only=True):
    if len(row) < 6:
        continue
    code = clean_code(row[0])
    desc = str(row[1]).strip() if row[1] else ''
    price = clean_price(row[5])
    if code and price and price > 0 and not code.startswith('HERMLE'):
        all_products.append({
            'code': code,
            'description': desc,
            'price': price,
            'currency': 'USD',
            'brand': 'HERMLE',
            'source': 'Hermle-2025'
        })
wb.close()
print(f'  Hermle: {sum(1 for p in all_products if p["brand"]=="HERMLE")} products')

# ===== 8. Biotech Rabbit (Price List O - EUR) =====
print('Loading Biotech Rabbit...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'Price List O 2025-03 EUR_Distributor.xlsx'), read_only=True, data_only=True)
ws = wb['EUR Distributor Prices 2025']
for row in ws.iter_rows(min_row=3, values_only=True):
    if len(row) < 8:
        continue
    code = clean_code(row[0])
    product_name = str(row[1]).strip() if row[1] else ''
    price = clean_price(row[7])
    if code and price and price > 0 and len(code) > 2:
        all_products.append({
            'code': code,
            'description': product_name,
            'price': price,
            'currency': 'EUR',
            'brand': 'BIOTECH RABBIT',
            'source': 'BiotechRabbit-EUR'
        })
wb.close()
print(f'  Biotech Rabbit: {sum(1 for p in all_products if p["brand"]=="BIOTECH RABBIT")} products')

# ===== 9. Thistle / Cleaver Scientific =====
print('Loading Thistle/Cleaver...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, '2025 Version 1 Gel and Chemi Price List - Cleaver Scientific in USD.xlsx'), read_only=True, data_only=True)
ws = wb['Raw Data for ERP Import']
for row in ws.iter_rows(min_row=2, values_only=True):
    code = clean_code(row[0])
    price = clean_price(row[1])
    desc = str(row[2]).strip() if len(row) > 2 and row[2] else ''
    if code and price and price > 0:
        all_products.append({
            'code': code,
            'description': desc,
            'price': price,
            'currency': 'USD',
            'brand': 'THISTLE',
            'source': 'Cleaver-GelChemi'
        })
wb.close()

wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'ROW Cleaver Brand Price List in USD 2025.xlsx'), read_only=True, data_only=True)
if 'Standard Products 2025' in wb.sheetnames:
    ws = wb['Standard Products 2025']
    for row in ws.iter_rows(min_row=7, values_only=True):
        if len(row) < 6:
            continue
        code = clean_code(row[3])
        desc = str(row[4]).strip() if row[4] else ''
        price = clean_price(row[5])
        if code and price and price > 0:
            all_products.append({
                'code': code,
                'description': desc,
                'price': price,
                'currency': 'USD',
                'brand': 'THISTLE',
                'source': 'Cleaver-ROW'
            })
wb.close()
print(f'  Thistle/Cleaver: {sum(1 for p in all_products if p["brand"]=="THISTLE")} products')

# ===== 10. Bio-Rad =====
print('Loading Bio-Rad...')
wb = openpyxl.load_workbook(os.path.join(PRICE_DIR, 'Formulario Pedido de Venda IA Order Vision - SINTESE.xlsx'), read_only=True, data_only=True)
# Sheet name has encoding issue - find it by index or partial match
price_sheet = [s for s in wb.sheetnames if 'Pre' in s or 'reco' in s.lower() or 'pre' in s.lower()]
ws = wb[price_sheet[0]] if price_sheet else wb[wb.sheetnames[2]]
for row in ws.iter_rows(min_row=2, values_only=True):
    if len(row) < 4:
        continue
    code = clean_code(row[1])
    desc = str(row[2]).strip() if row[2] else ''
    price = clean_price(row[3])
    if code and price and price > 0:
        all_products.append({
            'code': code,
            'description': desc,
            'price': price,
            'currency': 'BRL',
            'brand': 'BIO-RAD',
            'source': 'BioRad-Formulario'
        })
wb.close()
print(f'  Bio-Rad: {sum(1 for p in all_products if p["brand"]=="BIO-RAD")} products')

# ===== Deduplicate =====
seen = {}
deduped = []
for p in all_products:
    key = f"{p['code']}|{p['brand']}"
    if key not in seen:
        seen[key] = len(deduped)
        deduped.append(p)
    elif 'updated' in p['source'].lower() or 'instrument' in p['source'].lower():
        deduped[seen[key]] = p
all_products = deduped

# Save
with open('data/price-index.json', 'w', encoding='utf-8') as f:
    json.dump(all_products, f, ensure_ascii=False)

print(f'\n=== TOTAL: {len(all_products)} products indexed ===')
print('By brand:')
brands = {}
for p in all_products:
    brands[p['brand']] = brands.get(p['brand'], 0) + 1
for brand, count in sorted(brands.items(), key=lambda x: -x[1]):
    print(f'  {brand}: {count}')
